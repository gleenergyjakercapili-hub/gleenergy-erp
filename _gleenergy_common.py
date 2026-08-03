"""
Shared code for the Gleenergy backends (app.py = SQLite, app_postgres.py = Postgres)
====================================================================================

Both backends implement the SAME four storage methods over different databases.
Everything that is NOT database-specific lives here so there is ONE copy to fix:

  - request body shapes (Pydantic models)
  - API-key protection for the /api/* routes
  - email sending  (/api/send-email)
  - serving the app (index + static files in public/)

SMS was REMOVED (2026-08): follow-up outreach is email-only (SMS shifted to
the Meta Ads / Messenger space), so the Semaphore integration and its
subscription cost are gone — /api/send-sms no longer exists.

Each backend just defines its own `_conn()` + the four storage routes, then calls
`install_shared(app)` to attach all of the above.

-----------------------------------------------------------------------
SECURITY: server-side sign-in sessions (always on)
-----------------------------------------------------------------------
Every /api/* call (except health + the auth endpoints themselves) requires a
signed-in session:

  - POST /api/auth/login  {email, password}  verifies the password ON THE
    SERVER against the stored PBKDF2 hash in p2:employees and sets an
    HttpOnly session cookie (7 days, sliding). Failed attempts are
    rate-limited per IP + account.
  - GET  /api/auth/me     tells the browser whether its cookie is still valid.
  - POST /api/auth/logout deletes the session and clears the cookie.

The session rows live in a `sessions` table next to the kv table (each backend
supplies its own accessors via install_shared(app, auth_db=...)). Only the
SHA-256 of the token is stored, so the database alone can't forge a cookie.

Sessions are the ONLY credential. The old X-API-Key bypass was removed after
the key value repeatedly leaked into the served page — server-to-server jobs
sign in through /api/auth/login like a person. GLEENERGY_API_KEY now only
switches CORS into locked-down mode; its value grants no access. Cross-origin
access stays off unless GLEENERGY_CORS_ORIGINS lists origins.
"""

import os
import json
import ssl
import time
import hmac
import base64
import hashlib
import secrets
import datetime
import smtplib
import urllib.parse
from typing import List
from email.message import EmailMessage

from pydantic import BaseModel
from fastapi import Request
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from starlette.concurrency import run_in_threadpool

BASE_DIR = os.path.dirname(__file__)
PUBLIC_DIR = os.path.join(BASE_DIR, "public")
INDEX_FILE = os.path.join(PUBLIC_DIR, "index.html")
# The email config can live in three places; the first one that loads wins:
#   config/email_config.json        — office PC / local layout (config/ is gitignored)
#   /etc/secrets/email_config.json  — a Render "Secret File" (Render's mount point)
#   email_config.json               — Render also copies secret files into the app root
# This is what lets the cloud service get its SMTP credentials via the Render
# dashboard without the secret ever touching GitHub.
EMAIL_CONFIG_CANDIDATES = (
    os.path.join(BASE_DIR, "config", "email_config.json"),
    "/etc/secrets/email_config.json",
    os.path.join(BASE_DIR, "email_config.json"),
)

# Secret for server-to-server scripts/backups (X-API-Key header). Browsers use
# session cookies instead; this is never injected into the served page.
API_KEY = os.environ.get("GLEENERGY_API_KEY", "").strip()

# Paths reachable without a session (the auth endpoints must be, or nobody
# could ever sign in).
_OPEN_API_PATHS = {"/api/health", "/api/auth/login", "/api/auth/logout", "/api/auth/me"}

# ----------------------------------------------------------------------
# Sign-in sessions
# ----------------------------------------------------------------------
SESSION_COOKIE = "glee_sess"
SESSION_TTL = 7 * 24 * 3600          # 7 days, sliding (renewed on activity)
_RENEW_AFTER = 12 * 3600             # extend at most every ~12 hours

# In-memory cache so chatty storage traffic doesn't hit the sessions table on
# every request (single-process deployment; logout evicts explicitly).
_SESS_CACHE = {}
_SESS_CACHE_TTL = 60

# Failed-login tracker: {key: [timestamps]}. Per account+IP and per IP.
_LOGIN_FAILS = {}
_FAIL_WINDOW = 600                   # 10 minutes
_MAX_FAILS_ACCOUNT = 8
_MAX_FAILS_IP = 30


def _client_ip(request):
    fwd = request.headers.get("x-forwarded-for", "")
    if fwd:
        return fwd.split(",")[0].strip()
    return request.client.host if request.client else "?"


def _is_https(request):
    proto = request.headers.get("x-forwarded-proto", "").split(",")[0].strip().lower()
    return proto == "https" or request.url.scheme == "https"


def _fails_in_window(key, now):
    arr = [t for t in _LOGIN_FAILS.get(key, []) if now - t < _FAIL_WINDOW]
    if arr:
        _LOGIN_FAILS[key] = arr
    else:
        _LOGIN_FAILS.pop(key, None)
    return len(arr)


def _record_fail(key, now):
    _LOGIN_FAILS.setdefault(key, []).append(now)


def _pw_ok(password, emp):
    """Verify a password against one employee record — the exact mirror of the
    app's pwDerive/pwVerify (PBKDF2-SHA256, hex salt + hex 32-byte hash)."""
    pw = password or ""
    if emp.get("passHash"):
        try:
            salt = bytes.fromhex(emp.get("passSalt") or "")
            iters = int(emp.get("passIter") or 120000)
        except (ValueError, TypeError):
            return False
        digest = hashlib.pbkdf2_hmac("sha256", pw.encode("utf-8"), salt, iters).hex()
        return hmac.compare_digest(digest, emp["passHash"])
    if emp.get("pass") is not None:                      # legacy plaintext, pre-migration
        return hmac.compare_digest(str(emp["pass"]), pw)
    return False


def _emp_active(emp):
    status = emp.get("status") or "Regular"
    return status not in ("Resigned", "Inactive")


def _find_employee(kv_get, email):
    """Returns (employee_or_None, roster_missing). roster_missing=True only when
    p2:employees has never been written — a brand-new empty database."""
    raw = kv_get("p2:employees")
    if raw is None:
        return None, True
    try:
        emps = json.loads(raw)
    except (ValueError, TypeError):
        return None, False
    email = (email or "").strip().lower()
    for e in emps:
        if str(e.get("email", "")).strip().lower() == email:
            return e, False
    return None, False


def _token_hash(token):
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def _session_from_request(request, auth_db):
    """Look up (and slide) the session for this request's cookie. Sync — call
    via run_in_threadpool from async middleware."""
    token = request.cookies.get(SESSION_COOKIE) or ""
    if not token or not auth_db:
        return None
    th = _token_hash(token)
    now = time.time()
    cached = _SESS_CACHE.get(th)
    if cached and cached["expires_at"] > now and now - cached["checked"] < _SESS_CACHE_TTL:
        return cached
    row = auth_db["sess_get"](th)
    if not row or row["expires_at"] <= now:
        _SESS_CACHE.pop(th, None)
        return None
    if row["expires_at"] - now < SESSION_TTL - _RENEW_AFTER:      # sliding renewal
        auth_db["sess_touch"](th, now + SESSION_TTL)
        row["expires_at"] = now + SESSION_TTL
    entry = {"emp_id": row["emp_id"], "email": row["email"],
             "expires_at": row["expires_at"], "checked": now}
    _SESS_CACHE[th] = entry
    if len(_SESS_CACHE) > 500:                                     # keep the cache tiny
        for k, v in list(_SESS_CACHE.items()):
            if v["expires_at"] <= now or now - v["checked"] > _SESS_CACHE_TTL:
                _SESS_CACHE.pop(k, None)
    return entry


# NOTE: X-API-Key access was deliberately removed (2026-07-27). The key value
# leaked into the served page repeatedly and could not be rotated reliably, so
# sessions are the only credential. GLEENERGY_API_KEY now only switches CORS
# into locked-down mode; its value grants nothing.


# ----------------------------------------------------------------------
# Request body shapes. Declaring these lets the route handlers be plain
# `def` functions (see note in the backends): FastAPI parses + validates
# the JSON body for us and runs sync handlers in a worker thread, so the
# blocking database / SMTP / HTTP calls never freeze the event loop.
# ----------------------------------------------------------------------
class KeyBody(BaseModel):
    key: str = ""


class SetBody(BaseModel):
    key: str = ""
    value: str = ""


class ListBody(BaseModel):
    prefix: str = ""


class ChangesBody(BaseModel):
    since: str = ""        # opaque marker from the previous reply; "" = first call
    prefix: str = "p2:"


# ----------------------------------------------------------------------
# Change feed — shared marker codec + reply builder.
# Browsers poll /api/storage/changes with the marker from their previous
# reply and get back WHICH keys changed (keys only, never values) plus the
# current presence map. Both backends must speak the exact same marker
# format or a marker minted by one silently blinds a client of the other.
# ----------------------------------------------------------------------
UTC = datetime.timezone.utc

PRESENCE_KEY = "p2:presence"
FEED_EXCLUDE_KEYS = (PRESENCE_KEY,)                    # rides the reply separately —
                                                       # heartbeats must never set the marker
FEED_EXCLUDE_PREFIXES = ("p2:notifseen:", "p2:notifdismissed:", "p2:chatread:")
FEED_MAX_KEYS = 1000
FEED_LOOKBACK = datetime.timedelta(seconds=2)          # absorbs commit-order inversion / NTP steps


def parse_marker(since):
    """Marker -> aware UTC datetime, or None = 'send everything'. NEVER raises.
    An unreadable marker must fail OPEN (full resync, merely wasteful). The
    opposite failure — silently matching nothing — blinds the browser forever."""
    s = (since or "").strip()
    if not s:
        return None
    s = s.replace("T", " ")
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        dt = datetime.datetime.fromisoformat(s)
    except ValueError:
        return None
    if dt.tzinfo is None:                              # bare = UTC, like CURRENT_TIMESTAMP
        dt = dt.replace(tzinfo=UTC)
    return dt.astimezone(UTC)


def marker_text(dt):
    """The ONE wire format both backends emit: UTC, ISO, explicit Z."""
    return dt.astimezone(UTC).strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"


def feed_exclude_sql(ph):
    """WHERE-clause fragment excluding feed-noise keys. ph = '?' or '%s'."""
    parts, params = [], []
    if FEED_EXCLUDE_KEYS:
        parts.append("key NOT IN (" + ",".join([ph] * len(FEED_EXCLUDE_KEYS)) + ")")
        params.extend(FEED_EXCLUDE_KEYS)
    for p in FEED_EXCLUDE_PREFIXES:
        parts.append("key NOT LIKE " + ph)             # wildcard lives in the VALUE
        params.append(p + "%")
    return ((" AND " + " AND ".join(parts)) if parts else ""), params


def changes_reply(rows, since_echo, presence_raw):
    """rows: (key, updated_at); updated_at is str (SQLite) or aware datetime (PG)."""
    truncated = len(rows) > FEED_MAX_KEYS
    rows = rows[:FEED_MAX_KEYS]
    newest = parse_marker(since_echo)                  # seed with caller's marker: never walks backwards
    changes = []
    for key, at in rows:
        dt = at if isinstance(at, datetime.datetime) else parse_marker(at)
        if dt is None:
            continue
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=UTC)
        if newest is None or dt > newest:
            newest = dt
        changes.append({"key": key, "at": marker_text(dt)})
    return {
        "changes": changes,
        # No rows -> keep the caller's marker. Advancing it to "now" would step
        # over a write committing at this very instant.
        "marker": marker_text(newest) if newest else (since_echo or ""),
        "full": not since_echo,
        "truncated": truncated,
        "presence": presence_raw or "",                # raw JSON text of p2:presence
    }


class LoginBody(BaseModel):
    email: str = ""
    password: str = ""


class EmailAttachment(BaseModel):
    filename: str = ""
    b64: str = ""                       # base64 payload (data: URL prefix tolerated)
    mime: str = "application/octet-stream"


class EmailBody(BaseModel):
    to: str = ""
    subject: str = ""
    body: str = ""
    attachments: List[EmailAttachment] = []   # e.g. the proposal PDF


class XlsxBody(BaseModel):
    name: str = ""
    b64: str = ""          # base64 of the .xlsx/.xlsm (data: URL prefix tolerated)
    max_rows: int = 500
    max_cols: int = 50


# ----------------------------------------------------------------------
# Config loaders
# ----------------------------------------------------------------------
def _load_json(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


# ----------------------------------------------------------------------
# Email sender (plain function; the route below just wraps it).
# SMS/Semaphore was removed 2026-08 — outreach is email-only.
# ----------------------------------------------------------------------
def _send_email(body: EmailBody):
    to = (body.to or "").strip()
    subject = body.subject or "(no subject)"
    text = body.body or ""

    cfg = {}
    for _p in EMAIL_CONFIG_CANDIDATES:
        cfg = _load_json(_p)
        if cfg:
            break
    if not cfg.get("enabled"):
        return JSONResponse(
            {"ok": False, "error": "Email is not set up yet. Open email_config.json, fill in your details, and set enabled to true."},
            status_code=400,
        )
    if not to:
        return JSONResponse({"ok": False, "error": "This client has no email address."}, status_code=400)

    from_email = cfg.get("from_email") or cfg.get("smtp_user") or ""
    from_name = cfg.get("from_name") or "Gleenergy Renewables Company"
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = f"{from_name} <{from_email}>"
    msg["To"] = to
    reply_to = (cfg.get("reply_to") or "").strip()
    cc = (cfg.get("cc") or "").strip()
    bcc = (cfg.get("bcc") or "").strip()
    if reply_to:
        msg["Reply-To"] = reply_to     # where the client's reply goes
    if cc:
        msg["Cc"] = cc                 # visible copy
    if bcc:
        msg["Bcc"] = bcc               # private copy (stripped before sending, but still delivered)
    msg.set_content(text)

    # Attachments (e.g. the proposal PDF). Capped at 5 files / 15 MB total —
    # most mail providers reject anything larger anyway.
    total = 0
    for att in (body.attachments or [])[:5]:
        raw = (att.b64 or "").strip()
        if raw.lower().startswith("data:") and "," in raw:
            raw = raw.split(",", 1)[1]
        try:
            data = base64.b64decode(raw)
        except Exception:
            return JSONResponse({"ok": False, "error": f"Attachment {att.filename or '?'} is not valid base64."}, status_code=400)
        total += len(data)
        if total > 15 * 1024 * 1024:
            return JSONResponse({"ok": False, "error": "Attachments are over 15 MB in total — too large to email."}, status_code=400)
        mime = att.mime or "application/octet-stream"
        maintype, _, subtype = mime.partition("/")
        msg.add_attachment(data, maintype=maintype or "application", subtype=subtype or "octet-stream",
                           filename=att.filename or "attachment.bin")

    host = cfg.get("smtp_host", "smtp.gmail.com")
    port = int(cfg.get("smtp_port", 587))
    user = cfg.get("smtp_user", "")
    password = cfg.get("smtp_password", "")
    try:
        context = ssl.create_default_context()
        if port == 465:
            with smtplib.SMTP_SSL(host, port, context=context, timeout=25) as s:
                s.login(user, password)
                s.send_message(msg)
        else:
            with smtplib.SMTP(host, port, timeout=25) as s:
                s.ehlo()
                s.starttls(context=context)
                s.login(user, password)
                s.send_message(msg)
        return {"ok": True, "to": to}
    except Exception as e:
        # Don't leak internal SMTP detail to the caller; log it server-side.
        print(f"[send-email] failed: {e!r}")
        return JSONResponse(
            {"ok": False, "error": "Could not send the email. Check the server log and your email_config.json."},
            status_code=500,
        )


# ----------------------------------------------------------------------
# Excel (.xlsx / .xlsm) reader — returns each sheet (tab) as rows of text
# so the browser app can show an attached workbook read-only, tab by tab.
# Excel stays the source of truth: cached formula values are read, macros
# are NOT run.
# ----------------------------------------------------------------------
def _cell_to_str(v):
    import datetime
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, datetime.datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return v.strftime("%Y-%m-%d")
        return v.strftime("%Y-%m-%d %H:%M")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M")
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return str(round(v, 4))
    return str(v)


def _parse_xlsx(body: XlsxBody):
    import base64, io
    try:
        import openpyxl
    except Exception:
        return JSONResponse(
            {"ok": False, "error": "openpyxl is not installed on the server. Run: pip install openpyxl"},
            status_code=500,
        )
    raw = (body.b64 or "").strip()
    if raw.startswith("data:") and "," in raw:
        raw = raw.split(",", 1)[1]
    try:
        data = base64.b64decode(raw)
    except Exception:
        return JSONResponse({"ok": False, "error": "Could not decode the uploaded file."}, status_code=400)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"Not a readable Excel workbook: {e}"}, status_code=400)

    max_rows = max(1, min(int(body.max_rows or 500), 3000))
    max_cols = max(1, min(int(body.max_cols or 50), 120))
    sheets = []
    try:
        for ws in wb.worksheets:
            full_r = ws.max_row or 0
            full_c = ws.max_column or 0
            rmax, cmax = min(full_r, max_rows), min(full_c, max_cols)
            rows = []
            if rmax and cmax:
                for row in ws.iter_rows(min_row=1, max_row=rmax, max_col=cmax, values_only=True):
                    rows.append([_cell_to_str(c) for c in row])
            # trim trailing empty rows
            while rows and not any(rows[-1]):
                rows.pop()
            # trim trailing empty columns
            last = 0
            for x in rows:
                for ci in range(len(x) - 1, -1, -1):
                    if x[ci]:
                        if ci + 1 > last:
                            last = ci + 1
                        break
            rows = [x[:last] for x in rows]
            sheets.append({
                "name": ws.title,
                "hidden": ws.sheet_state != "visible",
                "rows": rows,
                "nrows": len(rows),
                "ncols": last,
                "truncated": full_r > max_rows or full_c > max_cols,
            })
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return {"ok": True, "name": body.name, "sheetCount": len(sheets), "sheets": sheets}


# ----------------------------------------------------------------------
# App-serving helpers
# ----------------------------------------------------------------------
def _serve_index():
    if not os.path.isfile(INDEX_FILE):
        return JSONResponse({"error": "index.html not found"}, status_code=404)
    # The page is served as-is. Browsers authenticate with their session cookie;
    # the API key is never injected (it used to be, which published it to anyone
    # who could load the page).
    return FileResponse(INDEX_FILE)


def _serve_static(path: str):
    if not path or path.startswith("api/"):
        return JSONResponse({"error": "not found"}, status_code=404)
    pub = os.path.abspath(PUBLIC_DIR)
    safe = os.path.abspath(os.path.join(pub, path))
    # block path traversal outside the public folder
    if not (safe == pub or safe.startswith(pub + os.sep)):
        return JSONResponse({"error": "not found"}, status_code=404)
    if os.path.isfile(safe):
        return FileResponse(safe)
    return JSONResponse({"error": "not found"}, status_code=404)


# ----------------------------------------------------------------------
# install_shared: attach CORS, the API-key guard, and all shared routes
# ----------------------------------------------------------------------
def install_shared(app, auth_db=None):
    """Attach everything that is common to both backends onto `app`.

    auth_db: dict of storage accessors the auth layer needs, supplied by each
    backend so this module stays database-agnostic:
      kv_get(key) -> value-string or None
      sess_get(token_hash) -> {"emp_id","email","expires_at"} or None
      sess_put(token_hash, emp_id, email, created_at, expires_at)
      sess_touch(token_hash, expires_at)
      sess_del(token_hash)
      sess_prune(now)          # delete expired rows
    """

    # --- CORS --------------------------------------------------------------
    if API_KEY:
        # Locked down: only origins you explicitly allow (usually none — the
        # app is served from the same origin, which never needs CORS).
        origins_env = os.environ.get("GLEENERGY_CORS_ORIGINS", "").strip()
        origins = [o.strip() for o in origins_env.split(",") if o.strip()]
        app.add_middleware(
            CORSMiddleware,
            allow_origins=origins,
            allow_methods=["*"],
            allow_headers=["*"],
        )
    else:
        # Open dev mode: permissive, same as the original starter.
        app.add_middleware(
            CORSMiddleware,
            allow_origins=["*"],
            allow_methods=["*"],
            allow_headers=["*"],
        )
        print("[gleenergy] NOTE: GLEENERGY_API_KEY is not set. Sign-in sessions guard "
              "the /api endpoints either way; the variable only switches CORS into "
              "locked-down mode for internet-facing deployments.")

    # --- Session guard for /api/* ------------------------------------------
    @app.middleware("http")
    async def _auth_guard(request: Request, call_next):
        if request.method != "OPTIONS":  # let CORS preflight through
            path = request.url.path
            if path.startswith("/api/"):
                # Cross-site POSTs are refused outright (SameSite=Lax already
                # keeps the cookie home; this also covers older browsers).
                if request.method == "POST":
                    origin = request.headers.get("origin", "")
                    if origin:
                        onet = urllib.parse.urlsplit(origin).netloc.lower()
                        host = request.headers.get("host", "").lower()
                        if onet and host and onet != host:
                            return JSONResponse(
                                {"ok": False, "error": "Cross-origin request refused."},
                                status_code=403,
                            )
                if path not in _OPEN_API_PATHS:
                    # Sign-in sessions are the ONLY way in. The X-API-Key bypass
                    # was removed after the key value leaked into the served page
                    # repeatedly — a key nobody can rotate reliably is not a
                    # credential. Server-to-server jobs sign in like a person.
                    authed = False
                    if auth_db is not None:
                        try:
                            sess = await run_in_threadpool(_session_from_request, request, auth_db)
                        except Exception as exc:                       # DB hiccup ≠ "not signed in"
                            print(f"[auth] session lookup failed: {exc!r}")
                            return JSONResponse(
                                {"ok": False, "error": "The database is briefly unavailable — please retry."},
                                status_code=503,
                            )
                        if sess:
                            authed = True
                            request.state.session = sess
                    if not authed and auth_db is None and not API_KEY:
                        authed = True   # no session backend wired and no key: open dev fallback
                    if not authed:
                        return JSONResponse(
                            {"ok": False, "error": "Not signed in."},
                            status_code=401,
                        )
        return await call_next(request)

    # --- Sign-in / session routes ------------------------------------------
    @app.post("/api/auth/login")
    def auth_login(body: LoginBody, request: Request):
        if auth_db is None:
            return JSONResponse({"ok": False, "error": "Sessions are not configured on this server."}, status_code=500)
        email = (body.email or "").strip().lower()
        password = body.password or ""
        if not email or not password:
            return JSONResponse({"ok": False, "error": "Email and password are required."}, status_code=400)
        now = time.time()
        ip = _client_ip(request)
        acct_key = ip + "|" + email
        if (_fails_in_window(acct_key, now) >= _MAX_FAILS_ACCOUNT
                or _fails_in_window("ip|" + ip, now) >= _MAX_FAILS_IP):
            return JSONResponse(
                {"ok": False, "error": "Too many sign-in attempts — please wait 10 minutes and try again."},
                status_code=429,
            )
        try:
            emp, roster_missing = _find_employee(auth_db["kv_get"], email)
        except Exception as exc:
            print(f"[auth] login roster read failed: {exc!r}")
            return JSONResponse(
                {"ok": False, "error": "The database is briefly unavailable — please try again in a moment."},
                status_code=503,
            )
        # Brand-new empty database: let the seed Super Admin in once so the app
        # can initialise itself. Never triggers once p2:employees exists.
        bootstrap = roster_missing and email == "ceo@solar" and password == "demo123"
        if not bootstrap and (emp is None or not _pw_ok(password, emp)):
            _record_fail(acct_key, now)
            _record_fail("ip|" + ip, now)
            return JSONResponse(
                {"ok": False, "error": "That email or password doesn't match. Please try again."},
                status_code=401,
            )
        if emp is not None and not _emp_active(emp):
            return JSONResponse(
                {"ok": False, "error": "This account is inactive — please contact your administrator."},
                status_code=403,
            )
        _LOGIN_FAILS.pop(acct_key, None)
        token = secrets.token_urlsafe(32)
        th = _token_hash(token)
        emp_id = emp["id"] if emp else "e0"
        try:
            auth_db["sess_prune"](now)
            auth_db["sess_put"](th, emp_id, email, now, now + SESSION_TTL)
        except Exception as exc:
            print(f"[auth] session write failed: {exc!r}")
            return JSONResponse(
                {"ok": False, "error": "The database is briefly unavailable — please try again in a moment."},
                status_code=503,
            )
        resp = JSONResponse({"ok": True, "empId": emp_id, "email": email})
        resp.set_cookie(SESSION_COOKIE, token, max_age=SESSION_TTL, httponly=True,
                        samesite="lax", secure=_is_https(request), path="/")
        return resp

    @app.get("/api/auth/me")
    def auth_me(request: Request):
        try:
            sess = _session_from_request(request, auth_db) if auth_db is not None else None
        except Exception as exc:
            print(f"[auth] /me lookup failed: {exc!r}")
            sess = None
        if not sess:
            return {"ok": False}
        return {"ok": True, "empId": sess["emp_id"], "email": sess["email"]}

    @app.post("/api/auth/logout")
    def auth_logout(request: Request):
        token = request.cookies.get(SESSION_COOKIE) or ""
        if token and auth_db is not None:
            th = _token_hash(token)
            try:
                auth_db["sess_del"](th)
            except Exception as exc:
                print(f"[auth] logout session delete failed: {exc!r}")
            finally:
                _SESS_CACHE.pop(th, None)
        resp = JSONResponse({"ok": True})
        resp.delete_cookie(SESSION_COOKIE, path="/")
        return resp

    # --- Shared routes -----------------------------------------------------
    @app.post("/api/send-email")
    def send_email(body: EmailBody):
        return _send_email(body)

    @app.post("/api/xlsx/parse")
    def xlsx_parse(body: XlsxBody):
        return _parse_xlsx(body)

    @app.get("/api/health")
    def health():
        return {"ok": True, "protected": True}   # sessions always guard the API

    @app.get("/")
    def index():
        return _serve_index()

    @app.get("/{path:path}")
    def static_files(path: str):
        return _serve_static(path)

    return app
